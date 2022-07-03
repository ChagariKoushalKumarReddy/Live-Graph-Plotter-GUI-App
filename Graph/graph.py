from tkinter import *
from tkinter import filedialog
import matplotlib

matplotlib.use("TkAgg")             # Using matplotlib as backend with Tkinter

# from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2TkAgg     # Importing the Canvas and the toolbar  
try:
    from matplotlib.backends.backend_tkagg import NavigationToolbar2TkAgg       # This block of code was written to counter the below error
except ImportError:
    from matplotlib.backends.backend_tkagg import NavigationToolbar2Tk as NavigationToolbar2TkAgg   # ImportError: cannot import name 'NavigationToolbar2TkAgg' from 'matplotlib.backends.backend_tkagg'

from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

from matplotlib.figure import Figure 
import matplotlib.animation as animation
import matplotlib.pyplot as plt
from matplotlib import style 
import csv
import time

from  openpyxl.workbook import Workbook         # openpyxl works only for excel files
from openpyxl import load_workbook              # for .csv files there is another module called CSV

style.use('bmh')        # Style used in graphs

window = Tk()

window.title("Graphing Calculator")
window.geometry("1000x700")

dir_xl = StringVar()       # Contains the directory of the excel file. StringVar is a string variable.StringVar() creates a "" empty string
timestamp = []          # Creating empty list timestamp to store the first column of values from the imported file
value = []              # Creating empty list value to store the second column of values from the imported file

def animate():
    a.clear()
    a.plot(timestamp,value)


def clicked(): 
# initial dir as '/' gives us the C: directory 
    global dir_xl
    window.filename = filedialog.askopenfilename(initialdir='/', title = 'Select A File', filetypes = ((".xlsx files","*.xlsx"),(".csv files","*.csv")))
    myLabel2.configure(text = window.filename)      # Configuring the label to display the file path 
    dir_xl = window.filename
    print(dir_xl)
    print(dir_xl[-4:])

# Button to import excel files

import_btn = Button(window, text = 'Import File', padx = 10, pady = 7,command = clicked, bd = 3)
import_btn.grid(row = 0 , column = 0)

# Labels to show the path of the file imported by the user

myLabel1 = Label(window, text = 'The path of the file is: ' , font = ('bold',12))
myLabel1.grid(row = 1 , column = 0,columnspan = 2,padx = 10, pady = 10 )
myLabel2 = Label(window, text = '', font= ('bold',10))
myLabel2.grid(row = 1, column = 2,padx = 10, pady = 10)

print("Koushal")
# print(dir_xl)


def plot():

    global timestamp
    global value
    global a

    timestamp = []              # making the lists 0 each time the button is clicked
    value = []
   
    if dir_xl[-4:]=='xlsx':                 # Checking whether the chosen file is excel or not
         # Extracting Excel data

        # wb = Workbook()     # Creating an workbook object

        # ws = wb.active      # Creating an active worksheet
        wb = load_workbook(dir_xl)  # Loading the spreadsheet from the given directory
        ws = wb.active
        column_1 = ws['A']         # This column_1 is a tuple
        column_2 = ws['B']
        for cell in column_1:
            timestamp.append(cell.value)        # taking the data from the 1st column and putting it in timestamp
            # print(cell.value)
        for cell in column_2:
            value.append(cell.value)            # taking the data from the 2nd column and putting it in timestamp
            # print(cell.value)
  
    if dir_xl[-3:]=='csv':                       # Checking whether the chosen file is .csv file or not
        with open(dir_xl,'r') as csv_file:
            csv_reader = csv.reader(csv_file)       # csv_reader is of the file type <class '_csv.reader'>
            print(type(csv_reader))
            for line in csv_reader:                 # line is of the type list
                timestamp.append( float (line[0]) )     # typecasting the elements present as float. Initially they were of string type 
                value.append( float (line[1]) )
                # print(line[0])
                # print(type(line))
    
    for i in range(len(timestamp)):     # i goes till len - 1
        a.clear()
        a.plot(timestamp[0:i+1],value[0:i+1])       # lst[0:i] gives till lst[i-1]
        show_plot()
        time.sleep(0.1)

    # print(timestamp)
    # print(value)

    # Pause button

    return


# Button to plot the graph

plot_btn = Button(window, text = 'Plot', padx = 20, pady = 8, command = plot, bd = 3)
plot_btn.grid(row = 0, column = 1)

f = Figure(figsize=(5,5), dpi = 100)
a = f.add_subplot(111)

# fig = plt.figure()
# plt.plot(timestamp,value)
def show_plot():
    canvas = FigureCanvasTkAgg(f, window)
    canvas.draw()
    canvas.get_tk_widget().grid(row = 2,column = 0,columnspan =3)

    toolbar = NavigationToolbar2TkAgg(canvas,window,pack_toolbar = False)   # By default it is getting packed so I turned it off
    toolbar.update()
    toolbar.grid(row = 3, column = 0, columnspan = 3)
    # canvas._tkcanvas.grid(row = 3, column = 0, columnspan = 3)

window.mainloop()